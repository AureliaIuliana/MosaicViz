from modules.preprocessing import load_samples_info, extract_variant_info, log_init, remove_empty_sheets
from modules.snv_analysis import SNV_analysis
from modules.indels_analysis import INDEL_analysis
import sys
import os
import logging
import openpyxl
from config import Config
import argparse
import warnings
warnings.simplefilter(action="ignore", category=FutureWarning) #ignore deprecated method

#------------------------------------------------------------------------------------------------------------------------------------------------------
#------------------------------------------------------------------------------------------------------------------------------------------------------
# HOW TO RUN the script: 
#       python3 MosaicViz.py -r PATH_REFERENCE -c PATH_CSV -d PATH_DIR_BAMS -o PATH_OUT_DIR
# HOW TO CHECK DETAILS of bam-readcount execution: 
#       bam-readcount -f hg19_simple_no_chr.fasta allbams/OPTI01_recal.bam 12:52200885-52200885 2>/dev/null

# ARGVS:
    #sys.argv[1] - path to bam_readcount executable file (-f bam-readcount parameter) - config.bam_readcount_arg
    #sys.argv[2] - path to reference sequence in fasta format - config.reference_arg
    #sys.argv[3] - MAPQ (-q bam-readcount parameter) - config.MAPQ_arg
    #sys.argv[4] - CSV file with informations about trio and variants (data_comma.csv) - config.csv_arg
    #sys.argv[5] - path to the directory containing all the BAMs to be investigated  - config.bams_directory_arg
    #sys.argv[6] - variant upstream and downstream analysis extension value (for INDEL analysis) - config.extension_arg
    #sys.argv[7] - lowerbound threshold to highlight proband/parents alternative allele percentage associated to potential mosaicism  - config.lower_threshold_arg
    #sys.argv[8] - upperbound threshold to highlight proband/parents alternative allele percentage associated to potential mosaicism  - config.upper_threshold_arg
    
# INFO DATA STRUCTURE: batch_MCD.csv -> 
#       SampleID,Project,samplesheet,Gene,coordinates,BAM/fastqname
#       1504V,p740,I,TUBA1A,12:49578981:G/A,1504V
#------------------------------------------------------------------------------------------------------------------------------------------------------

#logging 
main_logger = logging.getLogger("main")
main_logger.setLevel(logging.INFO)
bam_readcount_logger = logging.getLogger("bam-readcount execution")
bam_readcount_logger.setLevel(logging.INFO)
snv_analysis_logger = logging.getLogger("snv analysis")
snv_analysis_logger.setLevel(logging.INFO)
indel_analysis_logger = logging.getLogger("indel analysis")
indel_analysis_logger.setLevel(logging.INFO)

log_file = logging.FileHandler("../output/log.log", mode="w")
log_file.setLevel(logging.INFO)
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
log_file.setFormatter(formatter)

main_logger.addHandler(log_file)
bam_readcount_logger.addHandler(log_file)
snv_analysis_logger.addHandler(log_file)
indel_analysis_logger.addHandler(log_file)

# main 
def main():
    main_logger.info('Start processing')
    
    main_logger.info(f"Input arguments -b : {config.bam_readcount_arg}")
    main_logger.info(f"Input arguments -r : {config.reference_arg}")
    main_logger.info(f"Input arguments -c : {config.csv_arg}")
    main_logger.info(f"Input arguments -d : {config.bams_directory_arg}")
    main_logger.info(f"Input arguments -m : {config.mapq_arg}")
    main_logger.info(f"Input arguments -e : {config.extension_arg}")
    main_logger.info(f"Input arguments -l : {config.lower_threshold_arg}")
    main_logger.info(f"Input arguments -u : {config.upper_threshold_arg}")
    main_logger.info(f"Input arguments -o : {config.output_dir_arg}")
    
    workbook = openpyxl.Workbook()
    sheets = {
        'SNV': workbook.active,
        'DEL': workbook.create_sheet(title="Deletions"),
        'INS': workbook.create_sheet(title="Insertions"),
        'DELINS': workbook.create_sheet(title="DelIns")
        }
    sheets['SNV'].title = "SNVs"
    rows = {
        'SNV': 2,
        'DEL': 2,
        'INS': 2,
        'DELINS': 2
    }
    
    sample_info = load_samples_info(config.csv_arg, main_logger)
    for _, group_df in sample_info.groupby('trio'):
        trio_name = group_df.iloc[0, 1]
        chrom, posStart, posEnd_ref, posEnd_alt, ref, alt = extract_variant_info(group_df.iloc[0, 4])
        var = f"{ref}/{alt}"
        posStartExt = int(posStart) - int(config.extension_arg)
        posEndExt_ref = posEnd_ref + int(config.extension_arg)
        posEndExt_alt = posEnd_alt + int(config.extension_arg)
        gene = group_df.iloc[0, 3]
        bams = {
            'bam_file_I': f"{config.bams_directory_arg}/{group_df.iloc[0,5]}.bam",
            'bam_file_F': f"{config.bams_directory_arg}/{group_df.iloc[1,5]}.bam",
            'bam_file_M': f"{config.bams_directory_arg}/{group_df.iloc[2,5]}.bam"
        }
	# check bam existance
        for role, bam_path in bams.items():
            if not os.path.isfile(bam_path):
                error_msg = f"\n[ERROR] BAM file not found for trio '{trio_name}' ({role}): {bam_path}\nExecution aborted."
                print(error_msg)
                main_logger.error(error_msg)
                sys.exit(1)

        log_init(trio_name, bams, var, main_logger)
        print(f"Processing family: {trio_name}")
        
        # SNV Analysis
        if len(ref) == 1 and len(alt) == 1 and ref != "-" and alt != "-":
            main_logger.info("SNV nucleotide count analysis")
            print(f"target: {chrom}:{posStart} {var}")
            SNV_analysis(trio_name, bams, chrom, posStart, gene, var, ref, alt, rows["SNV"], sheets["SNV"], bam_readcount_logger, snv_analysis_logger, config)
            rows['SNV'] = rows['SNV'] + 5
        # DELETION Analysis
        elif ref != "-" and alt == "-":
            main_logger.info("DELETION nucleotide count analysis")
            print(f"target: {chrom}:{posStart} {var}")
            rows['DEL'] = INDEL_analysis(trio_name, bams, chrom, posStart, posStartExt, gene, posEndExt_ref, var, ref, alt, rows['DEL'], sheets['DEL'], "-", bam_readcount_logger, indel_analysis_logger, "DEL", config)
        # INSERTION Analysis
        elif ref == "-" and alt != "-":
            main_logger.info("INSERTION nucleotide count analysis")
            print(f"target: {chrom}:{posStart} {var}")
            rows['INS'] = INDEL_analysis(trio_name, bams, chrom, posStart, posStartExt, gene, posEndExt_alt, var, ref, alt, rows['INS'], sheets['INS'], "+", bam_readcount_logger, indel_analysis_logger, "INS", config)
        # DELINS Analysis
        else:
            main_logger.info("DELINS nucleotide count analysis")
            print(f"target: {chrom}:{posStart} {var}")
            SNV_analysis(trio_name, bams, chrom, posStart, gene, var, ref[0], alt[0], rows['DELINS'], sheets['DELINS'], bam_readcount_logger, snv_analysis_logger, config)
            rows['DELINS'] = rows['DELINS'] + 5

    output_excel_path = os.path.join(config.output_dir_arg, "results.xlsx")
    workbook.save(output_excel_path)
    main_logger.info(f"\nExcel saved at: {output_excel_path}")
    main_logger.info("\nEnd processing")


if __name__ == "__main__":
    
    parser = argparse.ArgumentParser(
                                    prog='MosaicViz', 
                                    usage='%(prog)s -r PATH_REFERENCE_FILE -c PATH_SAMPLE_INFO_CSV -d PATH_DIR_BAMS -o PATH_OUTPUT_DIR [OPTIONS]',
                                    description='''\
                                                MosaicViz is a python script that quantifies nucleotide read counts at target genomic positions using bam-readcount. It calculates nucleotides percentages across family trios and generates an Excel output with color-coded conditional formatting for rapid visual screening of potential mosaic variants''', 
                                    epilog='Example available at https://github.com/AureliaIuliana/MosaicViz',
                                    formatter_class=lambda prog: argparse.ArgumentDefaultsHelpFormatter(prog, max_help_position=40)
                                    )
    parser._optionals.title = "Arguments"
    parser.add_argument("-v", "--version", action='version', version='%(prog)s v.1.0')
    

    #arguments: mandatory 
    parser.add_argument("-b", "--bam-readcount", type=str, default="bam-readcount", help="path to bam-readcount executable file", metavar='')
    parser.add_argument("-r", "--reference", type=str, help="path to reference sequence in fasta format", metavar='')
    parser.add_argument("-c", "--csv", type=str, help="csv file with informations about trio identifiers and pedigree along with variant coordinates and bam files name", metavar='')
    parser.add_argument("-d", "--bams-directory", type=str, help="path to the directory containing all the BAMs to be investigated", metavar='')
    parser.add_argument("-o", "--output-dir", type=str, help="path to output dir", metavar='')
    
    #arguments: default  
    parser.add_argument("-m", "--mapq",  type=int, default=0, help="minimum mapping quality of reads used for counting", metavar='')
    parser.add_argument("-e", "--extension", type=int, default=1, help="variant upstream and downstream analysis extension value for INDEL analysis", metavar='')
    parser.add_argument("-l", "--lower-threshold", type=int, default=1, help="lower threshold value to highlight proband/parent alternative allele percentage associated to potential mosaicism", metavar='')
    parser.add_argument("-u", "--upper-threshold", type=int, default=40, help="upper threshold value to highlight proband/parent alternative allele percentage associated to potential mosaicism", metavar='')

    # Parse arguments
    args = parser.parse_args()

    # Create out dir if not exist 
    os.makedirs(args.output_dir, exist_ok=True)
    
    # Initialize the global config object
    config = Config(
        bam_readcount_arg = args.bam_readcount,
        reference_arg = args.reference,
        csv_arg = args.csv,
        bams_directory_arg = args.bams_directory,
        mapq_arg = args.mapq,
        extension_arg = args.extension,
        lower_threshold_arg = args.lower_threshold,
        upper_threshold_arg = args.upper_threshold,
        output_dir_arg = args.output_dir
    )
    
    main()
