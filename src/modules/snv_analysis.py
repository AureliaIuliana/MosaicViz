from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from modules.bamreadcount_exe import bam_readcount_execution, bam_readcount_data_extraction
from modules.indels_analysis import formatting_italic_gene
import sys
import pandas as pd

# GLOBAL VARIABLE
yellowFill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') #yellow
redFill = PatternFill(start_color='FA8C90', end_color='FA8C90', fill_type='solid') #red
light_greyFill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') #light grey
dark_greyFill = PatternFill(start_color='B3B3B3', end_color='B3B3B3', fill_type='solid') #dark grey
probandFill = PatternFill(start_color='FBF0B7', end_color='FBF0B7', fill_type='solid') # yellow
fatherFill = PatternFill(start_color='6ECFEA', end_color='6ECFEA', fill_type='solid') # blue
motherFill = PatternFill(start_color='FFB8C3', end_color='FFB8C3', fill_type='solid') # pink

def SNV_analysis(trio_name, bams, chrom, posStart, gene, var, ref, alt, sheet_row, sheet, logger, snv_logger, config):
    
    bamreadcount_coordinates_SNV = f'{chrom}:{posStart}-{posStart}'
    snv_logger.info(f"bamreadcount_coordinates_SNV: {bamreadcount_coordinates_SNV}")
        
    alleles_set = {"A", "C", "G", "T"}
    remaining_alleles = list(alleles_set - {ref, alt})
    allele1, allele2 = remaining_alleles
        
    #bam-readcount execution 
    df_p, max_col_p = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_I"], f'{bamreadcount_coordinates_SNV}', logger)
    df_f, max_col_f = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_F"], f'{bamreadcount_coordinates_SNV}', logger)    
    df_m, max_col_m = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_M"], f'{bamreadcount_coordinates_SNV}', logger)
    
    #data extraction from row data
    counts_p, percentages_p, depth_p = bam_readcount_data_extraction(df_p, "p", logger)
    counts_f, percentages_f, depth_f = bam_readcount_data_extraction(df_f, "f", logger)
    counts_m, percentages_m, depth_m = bam_readcount_data_extraction(df_m, "m", logger)
    
    #generation of the dataframe    
    df = pd.DataFrame({
            "individual": ["PROBAND", "FATHER", "MOTHER"],
            "projectID": [trio_name, trio_name, trio_name],
            "gene": [gene, gene, gene],
            "analyzed variant": [var, var, var],
            "chromosome": [chrom, chrom, chrom],
            "position": [posStart, posStart, posStart],
            "ref": [ref, ref, ref],
            "alt": [alt, alt, alt],
            "depth": [depth_p, depth_f, depth_m],
            "A": [counts_p["A_p"], counts_f["A_f"], counts_m["A_m"]],
            "C": [counts_p["C_p"], counts_f["C_f"], counts_m["C_m"]],
            "G": [counts_p["G_p"], counts_f["G_f"], counts_m["G_m"]],
            "T": [counts_p["T_p"], counts_f["T_f"], counts_m["T_m"]],
            "N": [counts_p["N_p"], counts_f["N_f"], counts_m["N_m"]],
            "A%": [percentages_p["A_perc_p"], percentages_f["A_perc_f"], percentages_m["A_perc_m"]],
            "C%": [percentages_p["C_perc_p"], percentages_f["C_perc_f"], percentages_m["C_perc_m"]],
            "G%": [percentages_p["G_perc_p"], percentages_f["G_perc_f"], percentages_m["G_perc_m"]],
            "T%": [percentages_p["T_perc_p"], percentages_f["T_perc_f"], percentages_m["T_perc_m"]],
            "N%": [percentages_p["N_perc_p"], percentages_f["N_perc_f"], percentages_m["N_perc_m"]],
            "BAMname": [bams["bam_file_I"], bams["bam_file_F"], bams["bam_file_M"]]
        })
    
    #adding a epty row at the end of the dataframe in order to have a empty row between trio data
    empty_row = {col: '' for col in df.columns}
    df = df.append(empty_row, ignore_index=True)
    
    #calculate the background noise 
    background_noise = max_background_noise_SNV(allele1, allele2, df)
    snv_logger.info(f"background_noise: {background_noise}")
    
    dataframe_to_excel(df, sheet)
        
    # sheet formatting 
    formatting_italic_gene(sheet) 
    formatting_sheet_SNV(light_greyFill, dark_greyFill, sheet, sheet_row) # per celle grigio chiaro e scuro 
    formatting_individual_column_sheets(sheet, probandFill, fatherFill, motherFill) # per celle giallo, blu, rosa
    conditional_formatting_proband_SNV(yellowFill, redFill, sheet, sheet_row, ref, config) # giallo per valori ref
    conditional_formatting_proband_SNV(yellowFill, redFill,  sheet, sheet_row, alt, config) # giallo per valori alt > 40 
    conditional_formatting_parents_SNV(yellowFill, redFill, sheet, sheet_row, ref, alt, background_noise, config)

def max_background_noise_SNV(allele1, allele2, df):
    noises = {}
    individuals = ["p", "f", "m"]
    rows = [0,1,2]
    for ind, r in zip(individuals,rows):
        pair = [1,2]
        alleles = [allele1, allele2]
        rows = [0,1,2]
        for i, allele in zip(pair, alleles):
                bn = f"bn{i}_{ind}"
                noises[bn] = df.loc[r, f"{allele}%"]
    background_noise = max(noises.values())
    return background_noise

#insert data from dataframe in excel sheet 
def dataframe_to_excel(df, sheet):
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
            
def conditional_formatting_proband_SNV(fillColor1, fillColor2, sheet, sheet_row, ref_or_alt, config):
    bases = ["A", "C", "G", "T"]
    sheet_cols = ["O", "P", "Q", "R"]
    for b, c in zip(bases, sheet_cols):
        lower_threshold = float(f"{config.lower_threshold_arg}")
        upper_threshold = float(f"{config.upper_threshold_arg}")
        if (ref_or_alt == b and sheet[f"{c}{sheet_row}"].value > lower_threshold and sheet[f"{c}{sheet_row}"].value < upper_threshold):
            sheet[f"{c}{sheet_row}"].fill = fillColor2
        elif (ref_or_alt == b): 
            sheet[f"{c}{sheet_row}"].fill = fillColor1

def conditional_formatting_parents_SNV(fillColor1, fillColor2, sheet, sheet_row, ref, alt, background_noise, config):
    bases = ["A", "C", "G", "T"]
    sheet_cols = ["O", "P", "Q", "R"]
    sheet_rows = [sheet_row+1, sheet_row+2]
    lower_threshold = float(f"{config.lower_threshold_arg}")
    upper_threshold = float(f"{config.upper_threshold_arg}")
    for r in sheet_rows:
        for b, c in zip(bases, sheet_cols):
            if (ref == b):
                sheet[f"{c}{r}"].fill = fillColor1
            if (alt == b and (sheet[f"{c}{r}"].value > background_noise)):
                if (sheet[f"{c}{r}"].value < lower_threshold):
                    sheet[f"{c}{r}"].fill = fillColor1
                elif (sheet[f"{c}{r}"].value > lower_threshold and sheet[f"{c}{r}"].value < upper_threshold): 
                    sheet[f"{c}{r}"].fill = fillColor2

def formatting_sheet_SNV(fillColor1, fillColor2, sheet, sheet_row):
    for r in [sheet_row, sheet_row+1, sheet_row+2]:
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
            sheet[f'{col}{r}'].fill = fillColor1 #light grey

    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T']:
        sheet[f'{col}{sheet_row-1}'].fill = fillColor2 #dark grey

def formatting_individual_column_sheets(sheet, fillColor1, fillColor2, fillColor3):
    colIndividual = sheet["A"]
    for cell in colIndividual:
        if (cell.value == "PROBAND"):
            cell.fill = fillColor1
        elif (cell.value == "FATHER"):
            cell.fill = fillColor2 
        elif (cell.value == "MOTHER"):
            cell.fill = fillColor3
