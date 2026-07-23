from openpyxl.styles import Font
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from modules.bamreadcount_exe import bam_readcount_execution
import sys
import pandas as pd

# GLOBAL VARIABLE
yellowFill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid') #yellow
redFill = PatternFill(start_color='FA8C90', end_color='FA8C90', fill_type='solid') #red
light_greyFill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') #light grey
dark_greyFill = PatternFill(start_color='B3B3B3', end_color='B3B3B3', fill_type='solid') #dark grey
probandFill = PatternFill(start_color='FBF0B7', end_color='FBF0B7', fill_type='solid') # yellow
fatherFill = PatternFill(start_color='6ECFEA', end_color='6ECFEA', fill_type='solid') # blue
motherFill = PatternFill(start_color='FFB8C3', end_color='FFB8C3', fill_type='solid') # pink

def INDEL_analysis(trio_name, bams, chrom, posStart, posStartExt, gene, posEndExt, var, ref, alt, sheet_row, sheet,  signINDEL, logger, indel_logger, indel_type, config):
    
    #check if delition or insertion:
    if indel_type == "DEL":
        ref_or_alt = ref
    elif indel_type == "INS":
        ref_or_alt = alt
        
    #coordinates 
    bamreadcount_extended_coordinates_INDEL = f"{chrom}:{posStartExt}-{posEndExt}"
    indel_logger.info(f"bamreadcount_extended_coordinates_INS: {bamreadcount_extended_coordinates_INDEL}")   
     
    # bam-readcount execution
    df_p, max_col_p = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_I"], f'{bamreadcount_extended_coordinates_INDEL}', logger)
    df_f, max_col_f = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_F"], f'{bamreadcount_extended_coordinates_INDEL}', logger)
    df_m, max_col_m = bam_readcount_execution(config.bam_readcount_arg, config.reference_arg, config.mapq_arg, bams["bam_file_M"], f'{bamreadcount_extended_coordinates_INDEL}', logger)
    
    #insert new data 
    ind_p = "PROBAND"
    ind_f = "FATHER"
    ind_m = "MOTHER"
    column_insertion_INDEL_dataframe(df_p, ind_p, trio_name, ref, alt, gene, chrom, bams["bam_file_I"], var)
    column_insertion_INDEL_dataframe(df_f, ind_f, trio_name, ref, alt, gene, chrom, bams["bam_file_F"], var)
    column_insertion_INDEL_dataframe(df_m, ind_m, trio_name, ref, alt, gene, chrom, bams["bam_file_M"], var)
    
    #concat all df trio 
    max_col = max(max_col_p, max_col_f, max_col_m)
    concat_df = pd.concat([df_p, df_f, df_m], axis=0)
    
    #esclude unnecessary columns 
    columns_to_exclude = ['Column_1', 'Column_5']
    subset_df = concat_df.drop(columns=columns_to_exclude)
    
    #add new columns 
    for base in ["A", "C", "G", "T", "N", "%INDEL"]:
        subset_df[f"{base}%"] = ""
    
    subset_df, matching_columns = esclude_non_target_indel(max_col, signINDEL, ref_or_alt, subset_df)

    new_col_name = f'{indel_type}:{signINDEL}{ref_or_alt}'
    if matching_columns:
        target_indel = f'{signINDEL}{ref_or_alt}'
        if matching_columns:
            subset_df[new_col_name] = subset_df.apply(
                lambda row: ' '.join(
                    val for val in row[matching_columns] if target_indel in str(val)
                ) if any(target_indel in str(val) for val in row[matching_columns]) else None,
                axis=1
            )
    
    columns_before_INDEL = ['individual', 'projectID', "gene", "analyzed variant", "chromosome", 'Column_2', 'Column_3', 'Column_4', 'Column_6', 'Column_7', 'Column_8', 'Column_9', 'Column_10', 'A%', 'C%', 'G%', 'T%', 'N%']
    column_INDEL = [col for col in subset_df.columns if col.startswith(f'{indel_type}:')]
    columns_name_after_INDEL = ['%INDEL', "BAMname"]
    all_column_names =  columns_before_INDEL + column_INDEL + columns_name_after_INDEL
    indel_logger.info(f"all_column_names: {all_column_names}")
    

    # Check for missing columns and handle them
    missing_cols = [col for col in all_column_names if col not in subset_df.columns]
    if missing_cols:
        indel_logger.warning(f"Warning: Missing columns in DataFrame: {missing_cols}")
        # Optionally handle missing columns here, e.g., by adding them with default values
        for col in missing_cols:
            subset_df[col] = "0"  # Or another placeholder
            
    subset_df = subset_df[all_column_names] 
    subset_df = subset_df.rename(columns={
                                'Column_2': 'position',
                                'Column_3': 'ref',
                                'Column_4': 'depth',
                                'Column_6': 'A',
                                'Column_7': 'C',
                                'Column_8': 'G',
                                'Column_9': 'T',
                                'Column_10': 'N'
                                })

    #extract count
    apply_base_extraction_count_value(subset_df)
    apply_indel_extraction_count_value(subset_df, new_col_name)
    #calculate percentage 
    base_percentage_calculation(subset_df, indel_logger)
    indel_percentage_calculation(subset_df, new_col_name, indel_logger)
    
    #add empy row between each trio 
    empty_row = {col: '' for col in subset_df.columns}
    subset_df = subset_df.append(empty_row, ignore_index=True)
    
    dataframe_to_excel(subset_df, sheet)
    
    #update shifted row value
    shift = update_shifted_row_value(signINDEL, sheet, sheet_row, ref, alt, config)
    
    # sheet formatting 
    formatting_italic_gene(sheet)
    formatting_header_sheet_INDEL(dark_greyFill, sheet)
    formatting_individual_column_sheets(sheet, probandFill, fatherFill, motherFill)
    formatting_pos_column_sheet_INDEL(sheet, sheet_row, posStart, dark_greyFill)
    formatting_sheet_INDEL(light_greyFill, sheet)
    conditional_formatting_ref_INDEL(redFill, yellowFill, sheet, sheet_row, ref)
    if new_col_name in subset_df.columns:
        conditional_formatting_alt_INDEL(redFill, yellowFill, sheet, sheet_row, config)
        
    sheet_row = sheet_row + shift
    return sheet_row
    
def column_insertion_INDEL_dataframe(df, individualName, trio_name, ref, alt, gene, chrom, bam_file_name, var):
    df.insert(0, "individual", individualName)
    df.insert(1, "projectID", trio_name)
    df.insert(2, "gene", gene)
    df.insert(3, "analyzed variant", var)
    df.insert(4, "chromosome", chrom)
    df.insert(5, "BAMname", bam_file_name)

def esclude_non_target_indel(max_col, signINDEL, ref_or_alt, subset_df): 
    matching_columns = []
    for i in range(11, max_col+1):
        col_name = f'Column_{i}'
        target_indel = f'{signINDEL}{ref_or_alt}'
        if any(target_indel == str(value).split(":")[0] for value in subset_df[col_name].values if isinstance(value, str)):
            matching_columns.append(col_name)
        else:
            subset_df = subset_df.drop(columns=col_name)
    return subset_df, matching_columns

def extract_count(row):
    if row is not None:
            return row.split(":")[1]  # Estrai la parte dopo ":"
    return None

def apply_base_extraction_count_value(df):
    for b in ["A", "C", "G", "T", "N"]:
        base_col = f"{b}" #f"{b}_count"
        df[base_col] = df[base_col].apply(extract_count)

def apply_indel_extraction_count_value(df, indel):
    target = indel.split(":")[1]
    if indel in df.columns:
        df[indel] = df[indel].apply(lambda row: extract_count(row) if target in str(row) else None)
    else:
        print(f"Column {indel} not found in DataFrame.")
    
def base_percentage_calculation(df, logger):
    for index, row in df.iterrows():
        depth_value = row.get("depth")
        
        if not depth_value:
            logger.warning(f"Warning: Depth is zero or missing for row {index}. Skipping calculation.")
            continue  
        
        depth = int(depth_value)
        base_columns = ["A", "C", "G", "T", "N"]  
        percent_columns = ["A%", "C%", "G%", "T%", "N%"]  
        
        for base_col, perc_col in zip(base_columns, percent_columns):
            base_count_value = row.get(base_col)
            row[perc_col] = round((int(base_count_value or 0) / depth) * 100, 2)

def indel_percentage_calculation(df, indel_col, logger):
    for index, row in df.iterrows():
        indelCountStr= row.get(indel_col)
        if indelCountStr is not None:
            indelCount = int(indelCountStr)
            depthStr = row.get("depth")
            depth = int(depthStr)
            row["%INDEL"] = (indelCount / depth) * 100
        else:
            logger.warning(f"Warning: indelCountString is None for row {index}")
            
def dataframe_to_excel(df, sheet):
    for row in dataframe_to_rows(df, index=False, header=True):
        sheet.append(row)
    
def update_shifted_row_value(sign,  sheet, sheet_row, ref, alt, config): 
    if sign == "-":
        shift = (len(ref) + (2 * int(config.extension_arg))) * 3 + 2    
    elif sign == "+":
        shift_INS_position("T", sheet, sheet_row)
        sheet[f'{"T"}{sheet_row}'].value = 0
        shift_INS_position("S", sheet, sheet_row)
        sheet[f'{"S"}{sheet_row}'].value = 0
        shift = (len(alt) + (2 * int(config.extension_arg))) * 3 + 2
    return shift
                
def formatting_italic_gene(sheet):
    for cell in sheet['C']:  
        cell.font = Font(italic=True)
            
def formatting_header_sheet_INDEL(fillColor1, sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row: 
            if "individual" in str(cell.value):
                for cell in row: 
                    cell.fill = fillColor1

def formatting_sheet_INDEL(fillColor1, sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell_value = str(cell.value)
            if "PROBAND" in cell_value or "FATHER" in cell_value or "MOTHER" in cell_value:
                for cell in row:
                    if cell.fill.start_color.index == "00000000" :
                        cell.fill = fillColor1
                break
    
def formatting_individual_column_sheets(sheet, fillColor1, fillColor2, fillColor3):
    colIndividual = sheet["A"]
    for cell in colIndividual:
        if (cell.value == "PROBAND"):
            cell.fill = fillColor1
        elif (cell.value == "FATHER"):
            cell.fill = fillColor2 
        elif (cell.value == "MOTHER"):
            cell.fill = fillColor3
                
def formatting_pos_column_sheet_INDEL(sheet, sheet_row, posStart, fillColor1):
    for i in range(sheet_row, sheet.max_row):
        if sheet[f'F{i}'].value == posStart:
            sheet[f'F{i}'].fill = fillColor1

def conditional_formatting_ref_INDEL(fillColor1, fillColor2, sheet, sheet_row, ref):
    bases = ["A", "C", "G", "T"]
    sheet_cols = ["N", "O", "P", "Q"]
    for i in range(sheet_row, sheet.max_row+1):
        for b,c in zip(bases, sheet_cols):
            if (b == sheet[f'G{i}'].value): # G=G
                sheet[f"{c}{i}"].fill = fillColor2 #N1

def conditional_formatting_alt_INDEL(fillColo1, fillColor2, sheet, sheet_row, config):
    for i in range(sheet_row, sheet.max_row+1):
        if sheet[f'S{i}'].value:
            cell_value = sheet[f'T{i}'].value
            cell_value_float = float(cell_value)
            lower_threshold = float(f"{config.lower_threshold_arg}")
            upper_threshold = float(f"{config.upper_threshold_arg}")
            background_noise = 1
            if (cell_value_float < upper_threshold and cell_value_float > lower_threshold):
                sheet[f'T{i}'].fill = fillColo1
            elif (cell_value_float > background_noise):
                sheet[f'T{i}'].fill = fillColor2

def shift_INS_position(excel_column, sheet, sheet_row):
    for i in range(sheet.max_row-1, sheet_row, -1):
        if sheet[f'{excel_column}{i-1}'].value:
            sheet[f'{excel_column}{i}'].value = sheet[f'{excel_column}{i-1}'].value
            sheet[f'{excel_column}{i-1}'].value = ''
